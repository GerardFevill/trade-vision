"""Export routes - CSV and Excel exports"""
import io
import csv
from datetime import datetime
from fastapi import APIRouter, Query, Request
from fastapi.responses import StreamingResponse
from slowapi import Limiter
from slowapi.util import get_remote_address
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from services import mt5_connector
from db import accounts_cache
from config.settings import settings
from config.logging import logger

router = APIRouter()
limiter = Limiter(key_func=get_remote_address)


def format_datetime(dt: datetime | str | None) -> str:
    """Format datetime for export"""
    if dt is None:
        return ""
    if isinstance(dt, str):
        return dt
    return dt.strftime("%Y-%m-%d %H:%M:%S")


@router.get("/export/trades/csv")
@limiter.limit("10/minute")
async def export_trades_csv(
    request: Request,
    days: int = Query(default=30, ge=1, le=365, description="Nombre de jours d'historique")
):
    """Export trade history as CSV"""
    logger.info("Exporting trades to CSV", days=days)
    trades = mt5_connector.get_history_trades(days)

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Ticket", "Symbol", "Type", "Volume",
        "Open Time", "Open Price", "Close Time", "Close Price",
        "Profit", "Commission", "Swap", "Comment"
    ])

    # Data
    for trade in trades:
        writer.writerow([
            trade.ticket,
            trade.symbol,
            trade.type,
            trade.volume,
            format_datetime(trade.open_time),
            trade.open_price,
            format_datetime(trade.close_time),
            trade.close_price or "",
            trade.profit,
            trade.commission,
            trade.swap,
            trade.comment
        ])

    output.seek(0)
    filename = f"trades_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/trades/excel")
@limiter.limit("10/minute")
async def export_trades_excel(
    request: Request,
    days: int = Query(default=30, ge=1, le=365, description="Nombre de jours d'historique")
):
    """Export trade history as Excel file"""
    logger.info("Exporting trades to Excel", days=days)
    trades = mt5_connector.get_history_trades(days)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header
    headers = [
        "Ticket", "Symbol", "Type", "Volume",
        "Open Time", "Open Price", "Close Time", "Close Price",
        "Profit", "Commission", "Swap", "Comment"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data
    for row, trade in enumerate(trades, 2):
        ws.cell(row=row, column=1, value=trade.ticket).border = thin_border
        ws.cell(row=row, column=2, value=trade.symbol).border = thin_border
        ws.cell(row=row, column=3, value=trade.type).border = thin_border
        ws.cell(row=row, column=4, value=trade.volume).border = thin_border
        ws.cell(row=row, column=5, value=format_datetime(trade.open_time)).border = thin_border
        ws.cell(row=row, column=6, value=trade.open_price).border = thin_border
        ws.cell(row=row, column=7, value=format_datetime(trade.close_time)).border = thin_border
        ws.cell(row=row, column=8, value=trade.close_price or "").border = thin_border
        cell = ws.cell(row=row, column=9, value=trade.profit)
        cell.border = thin_border
        if trade.profit > 0:
            cell.font = Font(color="008000")
        elif trade.profit < 0:
            cell.font = Font(color="FF0000")
        ws.cell(row=row, column=10, value=trade.commission).border = thin_border
        ws.cell(row=row, column=11, value=trade.swap).border = thin_border
        ws.cell(row=row, column=12, value=trade.comment).border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"trades_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/accounts/csv")
@limiter.limit("10/minute")
async def export_accounts_csv(request: Request):
    """Export all accounts summary as CSV"""
    logger.info("Exporting accounts to CSV")
    accounts = accounts_cache.load_accounts()
    if not accounts:
        accounts = mt5_connector.get_all_accounts_summary(use_cache=False)

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "ID", "Name", "Broker", "Server", "Balance", "Equity",
        "Profit", "Profit %", "Drawdown", "Trades", "Win Rate",
        "Currency", "Leverage", "Connected"
    ])

    # Data
    for acc in accounts:
        writer.writerow([
            acc.id,
            acc.name,
            acc.broker,
            acc.server,
            acc.balance,
            acc.equity,
            acc.profit,
            f"{acc.profit_percent:.2f}%",
            f"{acc.drawdown:.2f}%",
            acc.trades,
            f"{acc.win_rate:.2f}%",
            acc.currency,
            acc.leverage,
            "Yes" if acc.connected else "No"
        ])

    output.seek(0)
    filename = f"accounts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/accounts/excel")
@limiter.limit("10/minute")
async def export_accounts_excel(request: Request):
    """Export all accounts summary as Excel file"""
    logger.info("Exporting accounts to Excel")
    accounts = accounts_cache.load_accounts()
    if not accounts:
        accounts = mt5_connector.get_all_accounts_summary(use_cache=False)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header
    headers = [
        "ID", "Name", "Broker", "Server", "Balance", "Equity",
        "Profit", "Profit %", "Drawdown", "Trades", "Win Rate",
        "Currency", "Leverage", "Connected"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data
    for row, acc in enumerate(accounts, 2):
        ws.cell(row=row, column=1, value=acc.id).border = thin_border
        ws.cell(row=row, column=2, value=acc.name).border = thin_border
        ws.cell(row=row, column=3, value=acc.broker).border = thin_border
        ws.cell(row=row, column=4, value=acc.server).border = thin_border
        ws.cell(row=row, column=5, value=acc.balance).border = thin_border
        ws.cell(row=row, column=6, value=acc.equity).border = thin_border
        cell = ws.cell(row=row, column=7, value=acc.profit)
        cell.border = thin_border
        if acc.profit > 0:
            cell.font = Font(color="008000")
        elif acc.profit < 0:
            cell.font = Font(color="FF0000")
        ws.cell(row=row, column=8, value=f"{acc.profit_percent:.2f}%").border = thin_border
        ws.cell(row=row, column=9, value=f"{acc.drawdown:.2f}%").border = thin_border
        ws.cell(row=row, column=10, value=acc.trades).border = thin_border
        ws.cell(row=row, column=11, value=f"{acc.win_rate:.2f}%").border = thin_border
        ws.cell(row=row, column=12, value=acc.currency).border = thin_border
        ws.cell(row=row, column=13, value=acc.leverage).border = thin_border
        ws.cell(row=row, column=14, value="Yes" if acc.connected else "No").border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"accounts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/history/csv")
@limiter.limit("10/minute")
async def export_history_csv(
    request: Request,
    limit: int = Query(default=1000, ge=1, le=3600, description="Nombre de points")
):
    """Export balance/equity history as CSV"""
    logger.info("Exporting history to CSV", limit=limit)
    history = mt5_connector.get_history(limit)

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Timestamp", "Balance", "Equity", "Drawdown", "Drawdown %"
    ])

    # Data
    for point in history:
        writer.writerow([
            format_datetime(point.timestamp),
            point.balance,
            point.equity,
            point.drawdown,
            f"{point.drawdown_percent:.2f}%"
        ])

    output.seek(0)
    filename = f"history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
